# gui/reportes_window.py
import customtkinter as ctk
from tkinter import messagebox, filedialog, ttk
from datetime import datetime, timedelta
from tkcalendar import DateEntry
from database.db_manager import DatabaseManager
from core.session import obtener_sesion

class ReportesWindow:
    def __init__(self, parent):
        self.parent = parent

        # 🔹 Obtener sesión global
        self.session = obtener_sesion()

        # 🔹 Obtener la base de datos desde la sesión
        self.db = self.session.db

        # Colores del tema
        self.COLORES = {
            'card_bg': ("#FFFFFF", "#3a3a3a"),
            'primary': "#3498DB",
            'success': "#27AE60",
            'warning': "#F39C12",
            'danger': "#E74C3C",
        }

        # Variables
        self.fecha_inicio = datetime.now() - timedelta(days=30)
        self.fecha_fin = datetime.now()

        self._crear_interfaz()

    def _crear_interfaz(self):
        """Crea la interfaz principal"""
        # Container principal
        main_container = ctk.CTkFrame(self.parent, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=30, pady=30)

        # Header
        self._crear_header(main_container)

        # Panel de filtros
        self._crear_panel_filtros(main_container)

        # Tabs de reportes
        self._crear_tabs_reportes(main_container)

    def _crear_header(self, parent):
        """Crea el header con título"""
        header = ctk.CTkFrame(parent, fg_color="transparent")
        header.pack(fill="x", pady=(0, 20))

        # Título
        title_frame = ctk.CTkFrame(header, fg_color="transparent")
        title_frame.pack(side="left", fill="x", expand=True)

        ctk.CTkLabel(
            title_frame,
            text="📊 Reportes del Hotel",
            font=("Segoe UI", 28, "bold"),
            anchor="w"
        ).pack(anchor="w")

        ctk.CTkLabel(
            title_frame,
            text="Historial de reservas y habitaciones",
            font=("Segoe UI", 12),
            text_color=("#7F8C8D", "#95A5A6"),
            anchor="w"
        ).pack(anchor="w", pady=(5, 0))

        # Botón de exportación
        btn_frame = ctk.CTkFrame(header, fg_color="transparent")
        btn_frame.pack(side="right")

        btn_excel = ctk.CTkButton(
            btn_frame,
            text="📊 Exportar a Excel",
            command=self.exportar_excel,
            width=180,
            height=45,
            font=("Segoe UI", 12, "bold"),
            corner_radius=10,
            fg_color=self.COLORES['success'],
            hover_color="#229954"
        )
        btn_excel.pack(side="left", padx=5)

    def _crear_panel_filtros(self, parent):
        """Crea el panel de filtros de fecha"""
        filtros_container = ctk.CTkFrame(
            parent,
            fg_color=self.COLORES['card_bg'],
            corner_radius=15
        )
        filtros_container.pack(fill="x", pady=(0, 20))

        content = ctk.CTkFrame(filtros_container, fg_color="transparent")
        content.pack(fill="x", padx=20, pady=15)

        # Título
        ctk.CTkLabel(
            content,
            text="🗓️ Filtrar por Fecha",
            font=("Segoe UI", 14, "bold"),
            anchor="w"
        ).pack(side="left", padx=(0, 20))

        # Fecha inicio
        fecha_inicio_frame = ctk.CTkFrame(content, fg_color="transparent")
        fecha_inicio_frame.pack(side="left", padx=10)

        ctk.CTkLabel(
            fecha_inicio_frame,
            text="Desde:",
            font=("Segoe UI", 11, "bold")
        ).pack(side="left", padx=(0, 5))

        self.date_inicio = DateEntry(
            fecha_inicio_frame,
            width=15,
            background='#3498DB',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            font=("Segoe UI", 11),
            state='readonly',
            locale='es_ES'
        )
        self.date_inicio.pack(side="left")
        self.date_inicio.set_date(self.fecha_inicio)

        # Fecha fin
        fecha_fin_frame = ctk.CTkFrame(content, fg_color="transparent")
        fecha_fin_frame.pack(side="left", padx=10)

        ctk.CTkLabel(
            fecha_fin_frame,
            text="Hasta:",
            font=("Segoe UI", 11, "bold")
        ).pack(side="left", padx=(0, 5))

        self.date_fin = DateEntry(
            fecha_fin_frame,
            width=15,
            background='#3498DB',
            foreground='white',
            borderwidth=2,
            date_pattern='yyyy-mm-dd',
            font=("Segoe UI", 11),
            state='readonly',
            locale='es_ES'
        )
        self.date_fin.pack(side="left")
        self.date_fin.set_date(self.fecha_fin)

        # Botón aplicar filtros
        btn_aplicar = ctk.CTkButton(
            content,
            text="🔍 Aplicar",
            command=self._aplicar_filtros,
            width=100,
            height=35,
            font=("Segoe UI", 11, "bold"),
            corner_radius=8,
            fg_color=self.COLORES['primary'],
            hover_color="#2980B9"
        )
        btn_aplicar.pack(side="left", padx=10)

        # Botones rápidos
        btn_mes = ctk.CTkButton(
            content,
            text="Último Mes",
            command=lambda: self._filtro_rapido(30),
            width=100,
            height=35,
            font=("Segoe UI", 10),
            corner_radius=8,
            fg_color="transparent",
            border_width=2,
            border_color=("#BDC3C7", "#4A4A4A"),
            hover_color=("#ECF0F1", "#3A3A3A")
        )
        btn_mes.pack(side="left", padx=5)

        btn_trimestre = ctk.CTkButton(
            content,
            text="Último Trimestre",
            command=lambda: self._filtro_rapido(90),
            width=120,
            height=35,
            font=("Segoe UI", 10),
            corner_radius=8,
            fg_color="transparent",
            border_width=2,
            border_color=("#BDC3C7", "#4A4A4A"),
            hover_color=("#ECF0F1", "#3A3A3A")
        )
        btn_trimestre.pack(side="left", padx=5)

    def _crear_tabs_reportes(self, parent):
        """Crea las pestañas de reportes"""
        # Tabview
        self.tabview = ctk.CTkTabview(parent, corner_radius=15)
        self.tabview.pack(fill="both", expand=True)

        # Crear tabs
        self.tab_reservas = self.tabview.add("📅 Historial de Reservas")
        self.tab_habitaciones = self.tabview.add("🛏️ Historial de Habitaciones")

        # Cargar contenido inicial
        self._cargar_reporte_reservas()
        self._cargar_reporte_habitaciones()

    def _filtro_rapido(self, dias):
        """Aplica un filtro rápido de días"""
        self.date_inicio.set_date(datetime.now() - timedelta(days=dias))
        self.date_fin.set_date(datetime.now())
        self._aplicar_filtros()

    def _aplicar_filtros(self):
        """Aplica los filtros y recarga reportes"""
        self.fecha_inicio = self.date_inicio.get_date()
        self.fecha_fin = self.date_fin.get_date()

        if self.fecha_inicio > self.fecha_fin:
            messagebox.showerror(
                "Error",
                "La fecha de inicio no puede ser posterior a la fecha fin"
            )
            return

        # Recargar reportes
        self._cargar_reporte_reservas()
        self._cargar_reporte_habitaciones()

    def _cargar_reporte_reservas(self):
        """Carga el reporte de reservas"""
        # Limpiar contenido anterior
        for widget in self.tab_reservas.winfo_children():
            widget.destroy()

        # Container con scroll
        scroll_frame = ctk.CTkScrollableFrame(
            self.tab_reservas,
            fg_color="transparent"
        )
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Título
        ctk.CTkLabel(
            scroll_frame,
            text="Historial Completo de Reservas",
            font=("Segoe UI", 16, "bold"),
            anchor="w"
        ).pack(anchor="w", pady=(0, 15))

        # Obtener datos
        reservas = self.db.obtener_reporte_reservas(
            self.fecha_inicio.strftime("%Y-%m-%d"),
            self.fecha_fin.strftime("%Y-%m-%d")
        )

        if not reservas:
            ctk.CTkLabel(
                scroll_frame,
                text="No hay reservas en el período seleccionado",
                font=("Segoe UI", 12),
                text_color=("#7F8C8D", "#95A5A6")
            ).pack(pady=50)
            return

        # Crear tabla
        self._crear_tabla_reservas(scroll_frame, reservas)

    def _crear_tabla_reservas(self, parent, datos):
        """Crea una tabla con el historial de reservas"""
        # Frame para la tabla
        tabla_frame = ctk.CTkFrame(parent, fg_color=self.COLORES['card_bg'], corner_radius=10)
        tabla_frame.pack(fill="both", expand=True)

        # Crear Treeview
        style = ttk.Style()
        style.theme_use('clam')

        tree = ttk.Treeview(
            tabla_frame,
            columns=("ID", "Huésped", "Habitación", "Entrada", "Salida", "Días", "Total", "Estado"),
            show="headings",
            height=15
        )

        # Configurar columnas
        tree.heading("ID", text="ID")
        tree.heading("Huésped", text="Huésped")
        tree.heading("Habitación", text="Habitación")
        tree.heading("Entrada", text="Fecha Entrada")
        tree.heading("Salida", text="Fecha Salida")
        tree.heading("Días", text="Días")
        tree.heading("Total", text="Total")
        tree.heading("Estado", text="Estado")

        tree.column("ID", width=50, anchor="center")
        tree.column("Huésped", width=200)
        tree.column("Habitación", width=100, anchor="center")
        tree.column("Entrada", width=120, anchor="center")
        tree.column("Salida", width=120, anchor="center")
        tree.column("Días", width=70, anchor="center")
        tree.column("Total", width=100, anchor="e")
        tree.column("Estado", width=100, anchor="center")

        # Scrollbar
        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        # Insertar datos
        for reserva in datos:
            # Calcular días
            fecha_entrada = datetime.strptime(str(reserva[3]), "%Y-%m-%d")
            fecha_salida = datetime.strptime(str(reserva[4]), "%Y-%m-%d")
            dias = (fecha_salida - fecha_entrada).days

            tree.insert("", "end", values=(
                reserva[0],  # ID
                reserva[1],  # Huésped
                f"#{reserva[2]}",  # Habitación
                reserva[3],  # Entrada
                reserva[4],  # Salida
                dias,
                f"${reserva[5]:,.2f}",  # Total
                reserva[6].upper()  # Estado
            ))

        tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10, padx=(0, 10))

    def _cargar_reporte_habitaciones(self):
        """Carga el reporte de habitaciones"""
        # Limpiar contenido anterior
        for widget in self.tab_habitaciones.winfo_children():
            widget.destroy()

        # Container con scroll
        scroll_frame = ctk.CTkScrollableFrame(
            self.tab_habitaciones,
            fg_color="transparent"
        )
        scroll_frame.pack(fill="both", expand=True, padx=10, pady=10)

        # Título
        ctk.CTkLabel(
            scroll_frame,
            text="Historial de Uso y Limpieza de Habitaciones",
            font=("Segoe UI", 16, "bold"),
            anchor="w"
        ).pack(anchor="w", pady=(0, 15))

        # Obtener datos
        habitaciones = self.db.obtener_reporte_habitaciones(
            self.fecha_inicio.strftime("%Y-%m-%d"),
            self.fecha_fin.strftime("%Y-%m-%d")
        )

        if not habitaciones:
            ctk.CTkLabel(
                scroll_frame,
                text="No hay actividad de habitaciones en el período seleccionado",
                font=("Segoe UI", 12),
                text_color=("#7F8C8D", "#95A5A6")
            ).pack(pady=50)
            return

        # Crear tabla
        self._crear_tabla_habitaciones(scroll_frame, habitaciones)

    def _crear_tabla_habitaciones(self, parent, datos):
        """Crea una tabla con el historial de habitaciones"""
        # Frame para la tabla
        tabla_frame = ctk.CTkFrame(parent, fg_color=self.COLORES['card_bg'], corner_radius=10)
        tabla_frame.pack(fill="both", expand=True)

        # Crear Treeview
        tree = ttk.Treeview(
            tabla_frame,
            columns=("Habitación", "Tipo", "Evento", "Fecha", "Huésped", "Detalles"),
            show="headings",
            height=15
        )

        # Configurar columnas
        tree.heading("Habitación", text="Habitación")
        tree.heading("Tipo", text="Tipo")
        tree.heading("Evento", text="Evento")
        tree.heading("Fecha", text="Fecha")
        tree.heading("Huésped", text="Huésped")
        tree.heading("Detalles", text="Detalles")

        tree.column("Habitación", width=100, anchor="center")
        tree.column("Tipo", width=120, anchor="center")
        tree.column("Evento", width=150)
        tree.column("Fecha", width=120, anchor="center")
        tree.column("Huésped", width=200)
        tree.column("Detalles", width=200)

        # Scrollbar
        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)

        # Insertar datos
        for hab in datos:
            tree.insert("", "end", values=(
                f"#{hab[0]}",  # Habitación
                hab[1],  # Tipo
                hab[2],  # Evento
                hab[3],  # Fecha
                hab[4] if hab[4] else "-",  # Huésped
                hab[5] if hab[5] else "-"  # Detalles
            ))

        tree.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        scrollbar.pack(side="right", fill="y", pady=10, padx=(0, 10))

    def exportar_excel(self):
        """Exporta los reportes a Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

            # Diálogo para guardar
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"reporte_hotel_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )

            if not filename:
                return

            # Crear workbook
            wb = openpyxl.Workbook()

            # ===== HOJA 1: RESERVAS =====
            ws_reservas = wb.active
            ws_reservas.title = "Reservas"

            # Encabezado
            ws_reservas['A1'] = "HISTORIAL DE RESERVAS"
            ws_reservas['A1'].font = Font(size=16, bold=True)
            ws_reservas[
                'A2'] = f"Período: {self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')}"

            # Headers de tabla
            headers = ["ID", "Huésped", "Habitación", "Entrada", "Salida", "Días", "Total", "Estado"]
            for col, header in enumerate(headers, start=1):
                cell = ws_reservas.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="3498DB", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Datos de reservas
            reservas = self.db.obtener_reporte_reservas(
                self.fecha_inicio.strftime("%Y-%m-%d"),
                self.fecha_fin.strftime("%Y-%m-%d")
            )

            row = 5
            for reserva in reservas:
                fecha_entrada = datetime.strptime(str(reserva[3]), "%Y-%m-%d")
                fecha_salida = datetime.strptime(str(reserva[4]), "%Y-%m-%d")
                dias = (fecha_salida - fecha_entrada).days

                ws_reservas.cell(row, 1, reserva[0])
                ws_reservas.cell(row, 2, reserva[1])
                ws_reservas.cell(row, 3, f"#{reserva[2]}")
                ws_reservas.cell(row, 4, str(reserva[3]))
                ws_reservas.cell(row, 5, str(reserva[4]))
                ws_reservas.cell(row, 6, dias)
                ws_reservas.cell(row, 7, reserva[5])
                ws_reservas.cell(row, 7).number_format = '$#,##0.00'
                ws_reservas.cell(row, 8, reserva[6].upper())
                row += 1

            # Ajustar columnas
            ws_reservas.column_dimensions['A'].width = 8
            ws_reservas.column_dimensions['B'].width = 25
            ws_reservas.column_dimensions['C'].width = 12
            ws_reservas.column_dimensions['D'].width = 15
            ws_reservas.column_dimensions['E'].width = 15
            ws_reservas.column_dimensions['F'].width = 10
            ws_reservas.column_dimensions['G'].width = 15
            ws_reservas.column_dimensions['H'].width = 15

            # ===== HOJA 2: HABITACIONES =====
            ws_hab = wb.create_sheet("Habitaciones")

            # Encabezado
            ws_hab['A1'] = "HISTORIAL DE HABITACIONES"
            ws_hab['A1'].font = Font(size=16, bold=True)
            ws_hab['A2'] = f"Período: {self.fecha_inicio.strftime('%d/%m/%Y')} - {self.fecha_fin.strftime('%d/%m/%Y')}"

            # Headers de tabla
            headers_hab = ["Habitación", "Tipo", "Evento", "Fecha", "Huésped", "Detalles"]
            for col, header in enumerate(headers_hab, start=1):
                cell = ws_hab.cell(row=4, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="27AE60", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            # Datos de habitaciones
            habitaciones = self.db.obtener_reporte_habitaciones(
                self.fecha_inicio.strftime("%Y-%m-%d"),
                self.fecha_fin.strftime("%Y-%m-%d")
            )

            row = 5
            for hab in habitaciones:
                ws_hab.cell(row, 1, f"#{hab[0]}")
                ws_hab.cell(row, 2, hab[1])
                ws_hab.cell(row, 3, hab[2])
                ws_hab.cell(row, 4, str(hab[3]))
                ws_hab.cell(row, 5, hab[4] if hab[4] else "-")
                ws_hab.cell(row, 6, hab[5] if hab[5] else "-")
                row += 1

            # Ajustar columnas
            ws_hab.column_dimensions['A'].width = 12
            ws_hab.column_dimensions['B'].width = 15
            ws_hab.column_dimensions['C'].width = 20
            ws_hab.column_dimensions['D'].width = 15
            ws_hab.column_dimensions['E'].width = 25
            ws_hab.column_dimensions['F'].width = 30

            # Guardar
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Reporte exportado exitosamente a:\n{filename}")

        except ImportError:
            messagebox.showerror(
                "Error",
                "Para exportar a Excel necesitas instalar openpyxl:\npip install openpyxl"
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el reporte:\n{str(e)}")